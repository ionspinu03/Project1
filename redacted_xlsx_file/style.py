import openpyxl
from openpyxl.styles import Alignment, PatternFill


def insert_file_xlsx_from_redacted(path):
    # patch = "Investigare 11461 TURCANU SERGHEI.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    print("File is opened.")
    return sheet_obj, max_col, max_row, wb_obj

def delete_timestamp(sheet_obj):
    sheet_obj.delete_cols(6)
    print("Succeseful deleted rows.")
    print("Done")


def merge_cells(sheet_obj, max_col, max_row):

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column=3)
        if cell_obj.value.startswith("Depunere"):
            row = cell_obj.coordinate[1:]
            adress = "C" + row + ":" + "N" + row
            sheet_obj.merge_cells(adress)
            cell_obj.alignment = Alignment(horizontal='center')
            cell_obj.fill = PatternFill(patternType="solid", fgColor="32CD32")
        elif cell_obj.value.startswith("Plata"):
            row = cell_obj.coordinate[1:]
            adress = "C" + row + ":" + "N" + row
            sheet_obj.merge_cells(adress)
            cell_obj.alignment = Alignment(horizontal='center')
            cell_obj.fill = PatternFill(patternType="solid", fgColor="FF8C00")
    print("Succeseful merged cells and colored.")
    return sheet_obj

def cheange_type_value(sheet_obj, max_col, max_row):
    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column=3)
        cell_obj2 = sheet_obj.cell(row = i, column=6)
        cell_obj3 = sheet_obj.cell(row = i, column=7)
        cell_obj4 = sheet_obj.cell(row = i, column=14)
        cell_obj5 = sheet_obj.cell(row = i, column=13)
        if cell_obj.value.startswith("Depunere") or cell_obj.value.startswith("Plata"):
            pass
        else:
            cell_obj.value = int(cell_obj.value)
            cell_obj2.value = float(cell_obj2.value)
            cell_obj3.value = float(cell_obj3.value)
            cell_obj4.number_format = "HH:MM:SS"
            try:
                cell_obj5.value = (cell_obj5.value).replace("-","")
            except:pass
    for i in range(2,max_row+1):
        cell_obj = sheet_obj.cell(row = i, column=8)
        # print(cell_obj)
        try:
            if float(cell_obj.value) > 5000.00:
                # print(cell_obj)
                cell_obj.fill = PatternFill(patternType="solid", fgColor="FFC0CB")
        except:
            pass

def redacted_xlsx_file(path):
    sheet_obj, max_col, max_row, wb_obj = insert_file_xlsx_from_redacted(path)
    merge_cells(sheet_obj, max_col, max_row)
    delete_timestamp(sheet_obj)
    cheange_type_value(sheet_obj, max_col, max_row)
    wb_obj.save(path)


